from openpyxl.styles import Alignment, Font
from label_creator import LabelCreator

FONT_SIZE_10_BOLD = Font(size=10, bold=True)
FONT_SIZE_8 = Font(size=8)

class DataPopulator:
    def __init__(self, sheet):
        self.sheet = sheet

    def populate_workout_data(self, start_row, start_col, week_data, num_sets, set_width):
        for day_idx, day_data in enumerate(week_data):
            day_start_row = start_row + day_idx * 7 + 1
            for exercise_idx, exercise_data in enumerate(day_data['exercises']):
                self._populate_exercise_data(day_start_row + exercise_idx - 1, start_col + 3, exercise_data, num_sets, set_width)

    def _populate_exercise_data(self, start_row, start_col, exercise_data, num_sets, set_width):
        self._populate_exercise_name(start_row, start_col, exercise_data['name'])
        for set_idx, set_data in enumerate(exercise_data['sets']):
            self._populate_set_data(start_row, start_col + 2 + 1 + set_idx * (set_width + 1), set_data)

    def _populate_exercise_name(self, start_row, start_col, name):
        cell = self.sheet.cell(row=start_row, column=start_col)
        cell.value = name
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = FONT_SIZE_10_BOLD
        LabelCreator(self.sheet).add_border_to_label(start_row, start_col, start_row, start_col)

    def _populate_set_data(self, start_row, start_col, set_data):
        self._populate_cell(start_row, start_col, set_data.get('reps', ""), "right")
        self._populate_cell(start_row, start_col + 1, set_data.get('weight', ""), "right")
        self._populate_percentage_cell(start_row, start_col + 2, set_data.get('percentage_1rm', ""))
        self._populate_cell(start_row, start_col + 3, set_data.get('notes', ""), "left", FONT_SIZE_8)

    def _populate_cell(self, row, col, value, align, font=None):
        cell = self.sheet.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if font:
            cell.font = font

    def _populate_percentage_cell(self, row, col, value):
        cell = self.sheet.cell(row=row, column=col)
        if value is None or value == "":
            cell.value = ""
        else:
            cell.value = value / 100
            cell.number_format = '0%' if value == int(value) else '0.0%'
        cell.alignment = Alignment(horizontal="right", vertical="center")
