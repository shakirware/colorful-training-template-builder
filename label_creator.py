from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

BOLD_FONT = Font(bold=True)
FONT_SIZE_10_BOLD = Font(size=10, bold=True)
FONT_SIZE_10 = Font(size=10)
FONT_SIZE_8 = Font(size=8)
THIN_BORDER = Border(left=Side(style='thin', color='000000'),
                     right=Side(style='thin', color='000000'),
                     top=Side(style='thin', color='000000'),
                     bottom=Side(style='thin', color='000000'))

class LabelCreator:
    def __init__(self, sheet):
        self.sheet = sheet

    def add_border_to_label(self, start_row, start_col, end_row, end_col):
        self._apply_fill_and_border(start_row, start_col, end_row, end_col, None, THIN_BORDER)

    def _apply_fill_and_border(self, start_row, start_col, end_row, end_col, fill_color, border):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.sheet.cell(row=row, column=col)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = border

    def label_days(self, start_row, start_col, day_height, day_fill_color, exercise_fill_color, num_sets, set_width, grid_fill_color):
        days = [f"Day {i+1}" for i in range(7)]
        for i, day in enumerate(days):
            day_start_row = start_row + i * (day_height + 1)
            day_end_row = day_start_row + day_height - 1
            self._label_single_day(day, day_start_row, start_col + 1, day_end_row, day_fill_color)
            self._fill_exercises(day_start_row, day_end_row, start_col + 3, exercise_fill_color)
            self._fill_sets(day_start_row, day_end_row, start_col + 5 + 1, num_sets, set_width, grid_fill_color)

    def _label_single_day(self, day, day_start_row, day_start_col, day_end_row, day_fill_color):
        self.sheet.merge_cells(start_row=day_start_row, start_column=day_start_col, end_row=day_end_row, end_column=day_start_col)
        cell = self.sheet.cell(row=day_start_row, column=day_start_col)
        cell.value = day
        cell.fill = PatternFill(start_color=day_fill_color, end_color=day_fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = BOLD_FONT
        self.add_border_to_label(day_start_row, day_start_col, day_end_row, day_start_col)

    def _fill_exercises(self, start_row, end_row, start_col, exercise_fill_color):
        for row in range(start_row, end_row + 1):
            self._create_merged_cell(row, start_col, "", exercise_fill_color, FONT_SIZE_10_BOLD)

    def _create_merged_cell(self, row, col, value, fill_color, font):
        self.sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        cell = self.sheet.cell(row=row, column=col)
        cell.value = value
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = font
        self.add_border_to_label(row, col, row, col + 1)

    def _fill_sets(self, start_row, end_row, start_col, num_sets, set_width, grid_fill_color):
        for j in range(num_sets):
            set_start_col = start_col + j * (set_width + 1)
            self._apply_fill_and_border(start_row, set_start_col, end_row, set_start_col + set_width - 1, grid_fill_color, THIN_BORDER)

    def label_weeks(self, start_row, start_col, box_width, week_number, week_fill_color):
        week_label = f"Week {week_number}"
        start_column = start_col + 1
        end_column = start_col + box_width - 2
        self.sheet.merge_cells(start_row=start_row + 1, start_column=start_column, end_row=start_row + 1, end_column=end_column)
        cell = self.sheet.cell(row=start_row + 1, column=start_column)
        cell.value = week_label
        cell.fill = PatternFill(start_color=week_fill_color, end_color=week_fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = BOLD_FONT
        self.add_border_to_label(start_row + 1, start_column, start_row + 1, end_column)

    def label_sets(self, start_row, start_col, num_sets, set_fill_color, set_width, headers):
        for i in range(num_sets):
            set_start_col = start_col + 5 + i * (set_width + 1)
            self._label_single_set(start_row + 1, set_start_col, set_width, set_fill_color, i + 1)
            self._label_set_headers(start_row + 2, set_start_col, headers, set_fill_color)

    def _label_single_set(self, start_row, start_col, set_width, set_fill_color, set_number):
        self.sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + set_width - 1)
        cell = self.sheet.cell(row=start_row, column=start_col)
        cell.value = f"Set {set_number}"
        cell.fill = PatternFill(start_color=set_fill_color, end_color=set_fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = BOLD_FONT
        self.add_border_to_label(start_row, start_col, start_row, start_col + set_width - 1)

    def _label_set_headers(self, start_row, start_col, headers, set_fill_color):
        for j, header in enumerate(headers):
            header_cell = self.sheet.cell(row=start_row, column=start_col + j)
            header_cell.value = header
            header_cell.fill = PatternFill(start_color=set_fill_color, end_color=set_fill_color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.font = BOLD_FONT
            self.add_border_to_label(start_row, start_col + j, start_row, start_col + j)
