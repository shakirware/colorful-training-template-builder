from openpyxl.styles import PatternFill, Border, Side

THIN_BORDER = Border(left=Side(style='thin', color='000000'),
                     right=Side(style='thin', color='000000'),
                     top=Side(style='thin', color='000000'),
                     bottom=Side(style='thin', color='000000'))

class BoxCreator:
    def __init__(self, sheet):
        self.sheet = sheet

    def create_box(self, start_row, start_col, end_row, end_col, fill_color):
        self.fill_range(start_row, start_col, end_row, end_col, fill_color)
        self.add_border(start_row, start_col, end_row, end_col)

    def fill_range(self, start_row, start_col, end_row, end_col, fill_color):
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self.sheet.cell(row=row, column=col).fill = fill

    def add_border(self, start_row, start_col, end_row, end_col, border=THIN_BORDER):
        for col in range(start_col, end_col + 1):
            self.sheet.cell(row=start_row, column=col).border = Border(top=border.top)
            self.sheet.cell(row=end_row, column=col).border = Border(bottom=border.bottom)
        for row in range(start_row, end_row + 1):
            self.sheet.cell(row=row, column=start_col).border = Border(left=border.left)
            self.sheet.cell(row=row, column=end_col).border = Border(right=border.right)
