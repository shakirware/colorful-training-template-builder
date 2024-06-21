import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

class ColourfulTemplateGenerator:
    def __init__(self, workbook_name):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.workbook_name = workbook_name
        self.bold_font = Font(bold=True)  

    def create_box(self, start_row, start_col, end_row, end_col, fill_color):
        self.fill_range(start_row, start_col, end_row, end_col, fill_color)
        self.add_border(start_row, start_col, end_row, end_col)

    def fill_range(self, start_row, start_col, end_row, end_col, fill_color):
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self.sheet.cell(row=row, column=col).fill = fill

    def add_border(self, start_row, start_col, end_row, end_col, border_color='000000'):
        side = Side(style='thin', color=border_color)
        border = Border(left=side, right=side, top=side, bottom=side)

        for col in range(start_col, end_col + 1):
            self.sheet.cell(row=start_row, column=col).border = Border(top=side)
            self.sheet.cell(row=end_row, column=col).border = Border(bottom=side)
        for row in range(start_row, end_row + 1):
            self.sheet.cell(row=row, column=start_col).border = Border(left=side)
            self.sheet.cell(row=row, column=end_col).border = Border(right=side)

        self.sheet.cell(row=start_row, column=start_col).border = Border(top=side, left=side)
        self.sheet.cell(row=start_row, column=end_col).border = Border(top=side, right=side)
        self.sheet.cell(row=end_row, column=start_col).border = Border(bottom=side, left=side)
        self.sheet.cell(row=end_row, column=end_col).border = Border(bottom=side, right=side)

    def add_border_label(self, start_row, start_col, end_row, end_col):
        thin_side = Side(style='thin', color='000000')
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        self.fill_and_border(start_row, start_col, end_row, end_col, None, border)

    def fill_and_border(self, start_row, start_col, end_row, end_col, fill_color, border):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.sheet.cell(row=row, column=col)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = border

    def label_days(self, start_row, start_col, padding_top, padding_between, padding_side, set_padding_side, day_height, day_fill_color, num_sets, set_width, grid_fill_color):
        days = ["Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6", "Day 7"]
        for i, day in enumerate(days):
            day_start_row = start_row + padding_top + i * (day_height + padding_between)
            day_end_row = day_start_row + day_height - 1

            self.sheet.merge_cells(start_row=day_start_row, start_column=start_col + padding_side, end_row=day_end_row, end_column=start_col + padding_side)
            cell = self.sheet.cell(row=day_start_row, column=start_col + padding_side)
            cell.value = day
            cell.fill = PatternFill(start_color=day_fill_color, end_color=day_fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = self.bold_font  # Make text bold
            self.add_border_label(day_start_row, start_col + padding_side, day_end_row, start_col + padding_side)

            for j in range(num_sets):
                set_start_col = start_col + padding_side + 1 + set_padding_side + j * (set_width + padding_between)
                self.fill_and_border(day_start_row, set_start_col, day_end_row, set_start_col + set_width - 1, grid_fill_color, Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')))

    def label_weeks(self, start_row, start_col, box_width, week_number, padding_top, padding_side, week_fill_color):
        week_label = f"Week {week_number}"
        start_column = start_col + padding_side
        end_column = start_col + box_width - padding_side - 1
        self.sheet.merge_cells(start_row=start_row + padding_top, start_column=start_column, end_row=start_row + padding_top, end_column=end_column)
        cell = self.sheet.cell(row=start_row + padding_top, column=start_column)
        cell.value = week_label
        cell.fill = PatternFill(start_color=week_fill_color, end_color=week_fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = self.bold_font 
        self.add_border_label(start_row + padding_top, start_column, start_row + padding_top, end_column)

    def label_sets(self, start_row, start_col, num_sets, padding_top, padding_between, padding_side, set_fill_color, set_width, headers):
        for i in range(num_sets):
            set_start_col = start_col + padding_side + i * (set_width + padding_between)
            self.sheet.merge_cells(start_row=start_row + padding_top, start_column=set_start_col, end_row=start_row + padding_top, end_column=set_start_col + set_width - 1)
            cell = self.sheet.cell(row=start_row + padding_top, column=set_start_col)
            cell.value = f"Set {i + 1}"
            cell.fill = PatternFill(start_color=set_fill_color, end_color=set_fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = self.bold_font  
            self.add_border_label(start_row + padding_top, set_start_col, start_row + padding_top, set_start_col + set_width - 1)

            for j, header in enumerate(headers):
                header_cell = self.sheet.cell(row=start_row + padding_top + 1, column=set_start_col + j)
                header_cell.value = header
                header_cell.fill = PatternFill(start_color=set_fill_color, end_color=set_fill_color, fill_type="solid")
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                header_cell.font = self.bold_font  
                self.add_border_label(start_row + padding_top + 1, set_start_col + j, start_row + padding_top + 1, set_start_col + j)

    def create_consecutive_boxes(self, start_row, start_col, box_height, box_width, num_boxes, space_between, week_padding_top, week_padding_side, set_padding_top, set_padding_between, set_padding_side, num_sets, day_padding_top, day_padding_between, day_padding_side, day_height, fill_color, week_fill_color, set_fill_color, day_fill_color, grid_fill_color, set_width, headers):
        min_height = week_padding_top + 2 + set_padding_top + 2 + day_padding_top + (day_height * 7) + (day_padding_between * 6) + 1
        assert box_height >= min_height, f"The minimum height of the box must be {min_height}."

        min_width = week_padding_side + 1 + (num_sets * (set_width + set_padding_between)) + set_padding_side
        assert box_width >= min_width, f"The minimum width of the box must be {min_width}."

        for i in range(num_boxes):
            current_start_col = start_col + i * (box_width + space_between)
            self.create_box(start_row, current_start_col, start_row + box_height - 1, current_start_col + box_width - 1, fill_color)
            self.label_weeks(start_row, current_start_col, box_width, i + 1, week_padding_top, week_padding_side, week_fill_color)
            set_start_row = start_row + week_padding_top + 1
            self.label_sets(set_start_row, current_start_col + day_padding_side, num_sets, set_padding_top, set_padding_between, day_padding_side + set_padding_side, set_fill_color, set_width, headers)
            day_start_row = set_start_row + set_padding_top + 2
            self.label_days(day_start_row, current_start_col, day_padding_top, day_padding_between, day_padding_side, set_padding_side, day_height, day_fill_color, num_sets, set_width, grid_fill_color)

    def save_workbook(self):
        self.workbook.save(self.workbook_name)
        print(f"Workbook saved as {self.workbook_name}")

generator = ColourfulTemplateGenerator('workout_plan_new.xlsx')

generator.create_consecutive_boxes(
    start_row=2,
    start_col=2,
    box_height=28,
    box_width=15, 
    num_boxes=4,
    space_between=3,
    week_padding_top=1,
    week_padding_side=1,
    set_padding_top=1,
    set_padding_between=1,
    set_padding_side=1,
    num_sets=3,
    day_padding_top=1,
    day_padding_between=1,
    day_padding_side=1,
    day_height=2, 
    fill_color="FFCDAD",  
    week_fill_color="FFC099",  
    set_fill_color="FFB485",  
    day_fill_color="FFA770", 
    grid_fill_color="FF9A5C",
    set_width=3,
    headers=["Reps", "Sets", "%1RM"]
)

generator.save_workbook()
